# !/usr/bin/env python
# -*-coding:utf-8-*-
#from base64 import encode
import ctypes #win彩色字体
import sys
import os
import difflib
import datetime
from openpyxl import load_workbook
import docx
ERR_STR=f"""
        参考:
        {sys.argv[0]} <关键字>
            查找脚本所在当前目录下的excel文件与<关键字>匹配的项
        {sys.argv[0]} <关键字> -i <0.1-1>
            查找脚本所在当前目录下的excel文件与<关键字>匹配的项,匹配度要求10%到100%,写0.1到1之间的小数
        {sys.argv[0]} -s <关键字> -d {os.getcwd()}
            查找{os.getcwd()}目录下的excel文件与<关键字>匹配的项
        {sys.argv[0]} -s <关键字> -o
            查找脚本所在当前目录下的excel文件与<关键字>匹配的项,并且保存查找内容在csv文件中
        {sys.argv[0]} -s <关键字> -m 1
            查找脚本所在当前目录下的word文件与<关键字>匹配的项,1为mode1即word文档的意思
        >必选:
        
        [str:关键字]
        最后一个输入参数为关键字,用于匹配,和-s作用相同,只有在无-s参数时生效.

        -s [str:关键字]
            -s选择关键字,使用最后一行.

        >可选*:
        -m [int:1-2]
            -m选择模式,1为word,2为excel,默认为2,为了确保兼容性,只扫描docx和xlsx文件.

        *-h  None
            显示帮助,输入此项后只输出提示.

        *-d <str:目录或文件,?={os.getcwd()}>
            需要查找的目录/文件,默认当前目录及子目录所有的word和excel文件.(可拓展.xlsm,.xltx,.xltm,不过不推荐)

        *-i <float:数字,?=0.6>
            excel文件模糊匹配相似度,0.1-1,对应为10%到100%,越大要求相似度越高,默认为0.6.
            word文件只能强匹配,只要段落内包含关键字即可.

        *-o
            除了终端打印额外输出csv文件,文件名为alittlemc_out.csv.

        *-a <int:id>
            test测试.

    by alittlemc
    version:    1.2(2022-05-10=15:50:44)
"""

class bcolors:#兼容linux
    HEADER = '\033[0;37;40m\t'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    # '\033[0;37;40m\tHello World\033[0m'

STD_INPUT_HANDLE = -10
STD_OUTPUT_HANDLE = -11
STD_ERROR_HANDLE = -12
 
# 字体颜色定义 ,关键在于颜色编码，由2位十六进制组成，分别取0~f，前一位指的是背景色，后一位指的是字体色
#由于该函数的限制，应该是只有这16种，可以前景色与背景色组合。也可以几种颜色通过或运算组合，组合后还是在这16种颜色中
 
# Windows CMD命令行 字体颜色定义 text colors
FOREGROUND_BLACK = 0x00 # black.
FOREGROUND_DARKBLUE = 0x01 # dark blue.
FOREGROUND_DARKGREEN = 0x02 # dark green.
FOREGROUND_DARKSKYBLUE = 0x03 # dark skyblue.
FOREGROUND_DARKRED = 0x04 # dark red.
FOREGROUND_DARKPINK = 0x05 # dark pink.
FOREGROUND_DARKYELLOW = 0x06 # dark yellow.
FOREGROUND_DARKWHITE = 0x07 # dark white.
FOREGROUND_DARKGRAY = 0x08 # dark gray.
FOREGROUND_BLUE = 0x09 # blue.
FOREGROUND_GREEN = 0x0a # green.
FOREGROUND_SKYBLUE = 0x0b # skyblue.
FOREGROUND_RED = 0x0c # red.
FOREGROUND_PINK = 0x0d # pink.
FOREGROUND_YELLOW = 0x0e # yellow.
FOREGROUND_WHITE = 0x0f # white.
 
# Windows CMD命令行 背景颜色定义 background colors
BACKGROUND_BLUE = 0x10 # dark blue.
BACKGROUND_GREEN = 0x20 # dark green.
BACKGROUND_DARKSKYBLUE = 0x30 # dark skyblue.
BACKGROUND_DARKRED = 0x40 # dark red.
BACKGROUND_DARKPINK = 0x50 # dark pink.
BACKGROUND_DARKYELLOW = 0x60 # dark yellow.
BACKGROUND_DARKWHITE = 0x70 # dark white.
BACKGROUND_DARKGRAY = 0x80 # dark gray.
BACKGROUND_BLUE = 0x90 # blue.
BACKGROUND_GREEN = 0xa0 # green.
BACKGROUND_SKYBLUE = 0xb0 # skyblue.
BACKGROUND_RED = 0xc0 # red.
BACKGROUND_PINK = 0xd0 # pink.
BACKGROUND_YELLOW = 0xe0 # yellow.
BACKGROUND_WHITE = 0xf0 # white.
 
std_out_handle = ctypes.windll.kernel32.GetStdHandle(STD_OUTPUT_HANDLE)
 
def set_cmd_text_color(color, handle=std_out_handle):
    Bool = ctypes.windll.kernel32.SetConsoleTextAttribute(handle, color)
    return Bool
 
#重置色彩
def resetColor():
    set_cmd_text_color(FOREGROUND_RED | FOREGROUND_GREEN | FOREGROUND_BLUE)
 
#暗蓝色
#dark blue
def printDarkBlue(mess):
    set_cmd_text_color(FOREGROUND_DARKBLUE)
    sys.stdout.write(mess)
    resetColor()
 
#暗绿色
#dark green
def printDarkGreen(mess):
    set_cmd_text_color(FOREGROUND_DARKGREEN)
    sys.stdout.write(mess)
    resetColor()
 
#暗天蓝色
#dark sky blue
def printDarkSkyBlue(mess):
    set_cmd_text_color(FOREGROUND_DARKSKYBLUE)
    sys.stdout.write(mess)
    resetColor()
 
#暗红色
#dark red
def printDarkRed(mess):
    set_cmd_text_color(FOREGROUND_DARKRED)
    sys.stdout.write(mess)
    resetColor()
 
#暗粉红色
#dark pink
def printDarkPink(mess):
    set_cmd_text_color(FOREGROUND_DARKPINK)
    sys.stdout.write(mess)
    resetColor()
 
#暗黄色
#dark yellow
def printDarkYellow(mess):
    set_cmd_text_color(FOREGROUND_DARKYELLOW)
    sys.stdout.write(mess)
    resetColor()
 
#暗白色
#dark white
def printDarkWhite(mess):
    set_cmd_text_color(FOREGROUND_DARKWHITE)
    sys.stdout.write(mess)
    resetColor()
 
#暗灰色
#dark gray
def printDarkGray(mess):
    set_cmd_text_color(FOREGROUND_DARKGRAY)
    sys.stdout.write(mess)
    resetColor()
 
#蓝色
#blue
def printBlue(mess):
    set_cmd_text_color(FOREGROUND_BLUE)
    sys.stdout.write(mess)
    resetColor()
 
#绿色
#green
def printGreen(mess):
    set_cmd_text_color(FOREGROUND_GREEN)
    sys.stdout.write(mess)
    resetColor()
 
#天蓝色
#sky blue
def printSkyBlue(mess):
    set_cmd_text_color(FOREGROUND_SKYBLUE)
    sys.stdout.write(mess)
    resetColor()
 
#红色
#red
def printRed(mess):
    set_cmd_text_color(FOREGROUND_RED)
    sys.stdout.write(mess)
    resetColor()
 
#粉红色
#pink
def printPink(mess):
    set_cmd_text_color(FOREGROUND_PINK)
    sys.stdout.write(mess)
    resetColor()
 
#黄色
#yellow
def printYellow(mess):
    set_cmd_text_color(FOREGROUND_YELLOW)
    sys.stdout.write(mess)
    resetColor()
 
#白色
#white
def printWhite(mess):
    set_cmd_text_color(FOREGROUND_WHITE)
    sys.stdout.write(mess)
    resetColor()
 
#白底黑字
#white bkground and black text
def printWhiteBlack(mess):
    set_cmd_text_color(FOREGROUND_BLACK | BACKGROUND_WHITE)
    sys.stdout.write(mess)
    resetColor()
 
#白底黑字
#white bkground and black text
def printWhiteBlack_2(mess):
    set_cmd_text_color(0xf0)
    sys.stdout.write(mess)
    resetColor()
 
#黄底蓝字
#white bkground and black text
def printYellowRed(mess):
    set_cmd_text_color(BACKGROUND_YELLOW | FOREGROUND_RED)
    sys.stdout.write(mess)
    resetColor()


def printXLSXInfo(xlsxFile,search='',cutoff=0.6,out=''):
    # print(0)
    if len(out)>=4:
        # print(1)
        alm=[datetime.datetime.now().strftime('%m-%d %H:%M:%S'),'file:',"".join(xlsxFile.split()),'search:',search,'cutoff:',str(cutoff),'\n']
        # print(alm)
        savecvs(alm)
        
    wb = load_workbook(filename=xlsxFile)
    # print(xlsxFile)
    # errCounter = 0
    for x in wb:
        sheet_ranges = wb[x.title]
        for val in sheet_ranges.values:
            # print(val)
            val=list(filter(None,val))
            val=list(filter(lambda x: not str(x)[0]=='=',val))
            # print(val)
            # print(cutoff,type(cutoff))
            # print(val)
            # print(list(filter(lambda x: str(x),val)))
            # print(difflib.get_close_matches(search, list(filter(lambda x: not str(x).isnumeric(),val))))
            for pr in difflib.get_close_matches(search, list(filter(lambda x: not (type(x)==float or type(x)==int or str(x).isdigit() or str(x.replace('.','')).isdigit() ),val)),cutoff=float(cutoff)):
                printGreen('值>>'+pr+'\n>行>>'+str(val))
                print('\n')
                if len(out)>=4:
                    val.append('值:'+pr)
                    savecvs(val)
                    pass
                pass
            pass
        pass
    pass
         
def printWordInfo(WordFile,search='',out=''):
    # print(0)
    if len(out)>=4:
        alm=[datetime.datetime.now().strftime('%m-%d %H:%M:%S'),'file:',"".join(WordFile.split()),'search:',search,'\n']
        # print(alm)
        savecvs(alm)
    doc = docx.Document(WordFile)
    for para in doc.paragraphs:
        if search in para.text:
            if len(para.text)>=len(search):
                printSkyBlue('值>>'+search+f'\n>行>>'+para.text)
                print('\n')
                if len(out)>=4:
                    savecvs([para.text])
                    pass
                pass
            pass
        pass
    pass

# def printWordInfo(WordFile,search='',out=''):
#     #解压缩word
#     word = wc.Dispatch('Word.Application')
#     doc=word.Documents.Open(WordFile)
#     doc.SaveAs(WordFile+'load/',2)
#     doc.Close()
#     word.Quit()

#     pass

def addargv(str,argv=sys.argv,i=1,out=''):
    # print(str in argv)
    if ((str in argv) and (argv.index(str)+i<len(argv))):
        return argv[argv.index(str)+i]
    return out

def savecvs(date,name='alittlemc_out.csv'):
    with open(name,'a',encoding='utf-8-sig') as file:
        file.write(','.join(str(i) for i in date)+'\n')
        pass
    pass

def getfilename2(path,su):
    # input_template_All=[]
    # print(path,su)
    input_template_All_Path=[]
    for root, dirs, files in os.walk(path, topdown=False):
        # print(root,dirs)
        for name in files:
            # print(name)
            for su1 in su:
                # print(su1)
                if os.path.splitext(name)[1]==su1 and not name.startswith('~'):
                    input_template_All_Path.append(os.path.join(root, name))
                    pass
                pass
            pass
        pass
    return input_template_All_Path

def pac(xlsxFile):
    # print(0)
    wb = load_workbook(filename=xlsxFile)
    # print(xlsxFile)
    # errCounter = 0
    for x in wb:
        #全部的工作表
        sheet_ranges = wb[x.title]
        for val in sheet_ranges.values:
            # print(val)
            # val=list(filter(None,val))
            # val=list(filter(lambda x: not str(x)[0]=='=',val))

            savecvs(val[4:13],name='alittlemc_pac.csv')

            pass
        pass
    pass

if __name__ == '__main__':
    try:
        print(sys.argv)
        if '-h' in sys.argv:
            printDarkGray(ERR_STR)
        elif len(sys.argv)<=1:
            printRed('缺少参数\n'+ERR_STR)
        else:
            _m=int(addargv('-m',out='2'))
            _i=float(addargv('-i',out='0.6'))
            _d=addargv('-d',out=os.getcwd())
            
            if '-o' in sys.argv:
                _o='alittlemc_out.csv'
            else:
                _o=''

            _s=addargv('-s',out='')
            if len(_s)<=2:
                _s=sys.argv[len(sys.argv)-1]

            print(_i,_d,_o,_m,_s)
            # for file in os.listdir(_d):
            #     print(file)
            # print(getfilename2(_d,['.csv','.xls','.xlsx']))
            if _m==2:
                printGreen('excel模式')
                for fname in getfilename2(_d,['.xlsx']):
                    # print(_o)
                    if fname!=_o:
                        print('开始查找:'+fname)
                        printXLSXInfo(fname,search=_s,cutoff=_i,out=_o)
            # if _m==3:
            #     printGreen('pac模式')
            #     for fname in getfilename2(_d,['.xlsx']):
            #         # print(_o)
            #         if fname!=_o and '清单' in fname:
            #             print('开始查找:'+fname)
            #             pac(fname)
            else:
                printSkyBlue('word模式')
                for fname in getfilename2(_d,['.docx']):
                    # print(_o)
                    if fname!=_o:
                        print('开始查找:'+fname)
                        printWordInfo(fname,search=_s,out=_o)
    except Exception as e:
       printRed('#错误信息'+str(e)+'\n')